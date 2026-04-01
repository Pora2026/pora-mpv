import os
import re
import json
from datetime import date, datetime
from io import BytesIO

import openpyxl
from flask import Blueprint, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required

from app.config import INSTANCE_DIR
from app.extensions import db
from app.models import BusinessDay, ShiftRecord, ExpenseCategory, ExpenseEntry
from app.utils.dates import is_sunday, parse_ymd, iso, month_range


import_export_bp = Blueprint("import_export_bp", __name__)


def _owners():
    from app_owners import render_page, ensure_shifts, recalc_day_status, day_totals, range_series
    return render_page, ensure_shifts, recalc_day_status, day_totals, range_series


def _to_float_money(x) -> float:
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return 0.0
    s = s.replace("$", "").replace(" ", "")
    s = re.sub(r"[^0-9,.\-]", "", s)
    if not s:
        return 0.0

    if "," in s and "." in s:
        if s.rfind(".") > s.rfind(","):
            s = s.replace(",", "")
        else:
            s = s.replace(".", "").replace(",", ".")
        return float(s)

    if "," in s:
        if re.match(r"^-?\d{1,3}(,\d{3})+$", s):
            return float(s.replace(",", ""))
        return float(s.replace(",", "."))

    if "." in s:
        if re.match(r"^-?\d{1,3}(\.\d{3})+$", s):
            return float(s.replace(".", ""))

    return float(s)


def _norm_shift(s: str) -> str:
    s = (s or "").strip().lower()
    if s.startswith("ma"):
        return "Mañana"
    if s.startswith("ta"):
        return "Tarde"
    return (s.title() if s else "")


def _parse_date_cell(x):
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    s = str(x).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Fecha inválida en Excel: {x}")


def _find_header_map(ws):
    return 2, 3, 4, 5, 6


def build_export_data(d1: date, d2: date):
    _, ensure_shifts, recalc_day_status, day_totals, _ = _owners()

    days = (
        BusinessDay.query.filter(BusinessDay.day >= d1, BusinessDay.day <= d2)
        .order_by(BusinessDay.day.asc())
        .all()
    )

    out_days = []
    out_shifts = []
    out_expenses = []
    out_categories = []

    cats = ExpenseCategory.query.order_by(ExpenseCategory.kind.asc(), ExpenseCategory.name.asc()).all()
    for c in cats:
        out_categories.append(
            {"id": c.id, "kind": c.kind, "name": c.name, "created_at": c.created_at.isoformat() if c.created_at else None}
        )

    for d in days:
        if is_sunday(d.day):
            continue
        ensure_shifts(d)
        recalc_day_status(d)
        t = day_totals(d)

        out_days.append(
            {
                "date": d.day.isoformat(),
                "status": d.status,
                "note": d.note or "",
                "income": t["income"],
                "variable_expense": t["variable_expense"],
                "fixed_expense": t["fixed_expense"],
                "expense_total": t["expense_total"],
                "profit": t["profit"],
                "real_profit": d.real_profit,
            }
        )

        for s in d.shifts:
            out_shifts.append(
                {
                    "date": d.day.isoformat(),
                    "shift": s.shift,
                    "income": float(s.income or 0),
                    "note": s.note or "",
                    "is_closed": bool(s.is_closed),
                    "legacy_variable_expense_total": float(s.variable_expense_total or 0),
                    "legacy_fixed_expense_total": float(s.fixed_expense_total or 0),
                }
            )

        for e in d.expenses:
            out_expenses.append(
                {
                    "date": d.day.isoformat(),
                    "kind": e.kind,
                    "category_id": e.category_id,
                    "category_name": e.category.name if e.category else None,
                    "amount": float(e.amount or 0),
                    "note": e.note or "",
                    "created_at": e.created_at.isoformat() if e.created_at else None,
                }
            )

    return {
        "range": {"from": d1.isoformat(), "to": d2.isoformat()},
        "generated_at": datetime.utcnow().isoformat(),
        "days": out_days,
        "shifts": out_shifts,
        "expenses": out_expenses,
        "categories": out_categories,
    }


def export_to_excel(data: dict) -> BytesIO:
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    d1 = data["range"]["from"]
    d2 = data["range"]["to"]

    days = data["days"]
    total_income = sum(d["income"] for d in days)
    total_exp = sum(d["expense_total"] for d in days)
    total_profit = total_income - total_exp

    ws_sum.append(["Rango", f"{d1} a {d2}"])
    ws_sum.append(["Ingresos", total_income])
    ws_sum.append(["Gastos", total_exp])
    ws_sum.append(["Ganancia", total_profit])
    ws_sum.append(["Días (sin domingos)", len(days)])
    ws_sum.append([])
    ws_sum.append(["Nota", "Excel numérico (sin formato pesos). El backup real para reimport es el JSON."])

    ws_days = wb.create_sheet("Days")
    ws_days.append(
        ["Fecha", "Estado", "Nota", "Ingresos", "Gasto variable", "Gasto fijo", "Gasto total", "Ganancia", "Ganancia Real"]
    )
    for d in days:
        ws_days.append(
            [
                d["date"],
                d["status"],
                d["note"],
                d["income"],
                d["variable_expense"],
                d["fixed_expense"],
                d["expense_total"],
                d["profit"],
                d.get("real_profit"),
            ]
        )

    ws_exp = wb.create_sheet("Expenses")
    ws_exp.append(["Fecha", "Tipo", "Categoría", "Monto", "Nota", "Creado"])
    for e in data["expenses"]:
        ws_exp.append([e["date"], e["kind"], e.get("category_name"), e["amount"], e["note"], e.get("created_at")])

    ws_cat = wb.create_sheet("Categories")
    ws_cat.append(["ID", "Tipo", "Nombre", "Creado"])
    for c in data["categories"]:
        ws_cat.append([c["id"], c["kind"], c["name"], c.get("created_at")])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@import_export_bp.get("/export")
@login_required
def export_get():
    render_page, _, _, _, _ = _owners()

    today = date.today()
    d1, d2 = month_range(today)

    body = f"""
    <h1>Exportar / Backup</h1>

    <div class="card">
      <form method="get" action="/export/download">
        <div class="row-actions">
          <div class="field">
            <label>Desde</label>
            <input type="date" name="from" value="{iso(d1)}" />
          </div>
          <div class="field">
            <label>Hasta</label>
            <input type="date" name="to" value="{iso(d2)}" />
          </div>
          <div class="field">
            <label>Formato</label>
            <select name="fmt">
              <option value="json">JSON (reimportable / backup real)</option>
              <option value="xlsx">Excel (lectura humana)</option>
            </select>
          </div>
          <div style="min-width:180px;">
            <label>&nbsp;</label>
            <button class="btn primary" type="submit" style="width:100%;">Descargar</button>
          </div>
        </div>
        <p class="muted" style="margin-top:10px;">
          Recomendación: guardá el JSON siempre. El Excel es para mirar.
        </p>
      </form>
    </div>
    """
    return render_page(body, show_nav=True)


@import_export_bp.get("/export/download")
@login_required
def export_download():
    fmt = (request.args.get("fmt") or "json").strip().lower()
    d1s = (request.args.get("from") or "").strip()
    d2s = (request.args.get("to") or "").strip()

    if not d1s or not d2s:
        flash("Falta rango de fechas.", "error")
        return redirect(url_for("import_export_bp.export_get"))

    try:
        d1 = parse_ymd(d1s)
        d2 = parse_ymd(d2s)
    except ValueError:
        flash("Fechas inválidas.", "error")
        return redirect(url_for("import_export_bp.export_get"))

    if d1 > d2:
        d1, d2 = d2, d1

    data = build_export_data(d1, d2)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"owners_export_{d1.isoformat()}_{d2.isoformat()}_{stamp}"

    if fmt == "xlsx":
        bio = export_to_excel(data)
        return send_file(
            bio,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{base_name}.xlsx",
        )

    bio = BytesIO(json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8"))
    return send_file(bio, mimetype="application/json", as_attachment=True, download_name=f"{base_name}.json")


def import_balance_excel(filepath: str, sheet_names: list[str], mode: str = "skip") -> dict:
    _, ensure_shifts, recalc_day_status, _, _ = _owners()

    wb = openpyxl.load_workbook(filepath, data_only=True)
    imported = 0
    skipped = 0
    replaced = 0

    for sname in sheet_names:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        col_date, col_shift, col_income, col_var, col_fix = _find_header_map(ws)

        last_date = None
        for r in range(3, ws.max_row + 1):
            raw_date = ws.cell(r, col_date).value
            raw_shift = ws.cell(r, col_shift).value
            if raw_shift in (None, ""):
                continue

            if raw_date in (None, ""):
                if last_date is None:
                    continue
                d = last_date
            else:
                try:
                    d = _parse_date_cell(raw_date)
                    last_date = d
                except ValueError:
                    continue

            if is_sunday(d):
                continue

            shift = _norm_shift(str(raw_shift or ""))
            if shift not in ("Mañana", "Tarde"):
                continue

            income = _to_float_money(ws.cell(r, col_income).value)
            var_exp = _to_float_money(ws.cell(r, col_var).value)
            fix_exp = _to_float_money(ws.cell(r, col_fix).value)

            if income == 0 and var_exp == 0 and fix_exp == 0:
                continue

            bday = BusinessDay.query.filter_by(day=d).first()
            if not bday:
                bday = BusinessDay(day=d, note="", status="draft")
                db.session.add(bday)
                db.session.flush()
                ensure_shifts(bday)

            sr = ShiftRecord.query.filter_by(business_day_id=bday.id, shift=shift).first()
            if sr and mode == "skip":
                skipped += 1
                continue

            if not sr:
                sr = ShiftRecord(business_day=bday, shift=shift)
                db.session.add(sr)

            if sr.id and mode == "replace":
                replaced += 1
            else:
                imported += 1

            sr.income = income
            sr.variable_expense_total = var_exp
            sr.fixed_expense_total = fix_exp
            sr.is_closed = True
            recalc_day_status(bday)

    db.session.commit()
    return {"imported": imported, "replaced": replaced, "skipped": skipped}


def import_export_json(payload: dict, mode: str = "skip") -> dict:
    _, ensure_shifts, recalc_day_status, _, _ = _owners()

    imported = 0
    skipped = 0
    replaced = 0

    cats = payload.get("categories") or []
    for c in cats:
        kind = (c.get("kind") or "").strip().lower()
        name = (c.get("name") or "").strip()
        if kind not in ("fixed", "variable") or not name:
            continue
        ex = ExpenseCategory.query.filter_by(kind=kind, name=name).first()
        if not ex:
            db.session.add(ExpenseCategory(kind=kind, name=name))

    db.session.flush()

    def get_or_create_cat(kind: str, name: str):
        kind = (kind or "").strip().lower()
        name = re.sub(r"\s+", " ", (name or "").strip())
        if kind not in ("fixed", "variable") or not name:
            return None
        ex = ExpenseCategory.query.filter_by(kind=kind, name=name).first()
        if ex:
            return ex
        ex = ExpenseCategory(kind=kind, name=name)
        db.session.add(ex)
        db.session.flush()
        return ex

    days = payload.get("days") or []
    day_map = {}
    for d in days:
        ds = (d.get("date") or "").strip()
        if not ds:
            continue
        try:
            dd = parse_ymd(ds)
        except Exception:
            continue
        if is_sunday(dd):
            continue

        bday = BusinessDay.query.filter_by(day=dd).first()
        if bday and mode == "skip":
            skipped += 1
            day_map[ds] = bday
            continue

        if not bday:
            bday = BusinessDay(day=dd, note="", status="draft")
            db.session.add(bday)
            db.session.flush()
            ensure_shifts(bday)
            imported += 1
        else:
            replaced += 1

        bday.note = d.get("note") or ""
        bday.status = d.get("status") or "draft"
        bday.real_profit = d.get("real_profit", None)
        day_map[ds] = bday

        if mode == "replace":
            ExpenseEntry.query.filter_by(business_day_id=bday.id).delete()
            ShiftRecord.query.filter_by(business_day_id=bday.id).delete()
            db.session.flush()
            ensure_shifts(bday)

    db.session.flush()

    shifts = payload.get("shifts") or []
    for s in shifts:
        ds = (s.get("date") or "").strip()
        sh = (s.get("shift") or "").strip()
        if ds not in day_map:
            continue
        bday = day_map[ds]
        if sh not in ("Mañana", "Tarde"):
            continue

        sr = ShiftRecord.query.filter_by(business_day_id=bday.id, shift=sh).first()
        if not sr:
            sr = ShiftRecord(business_day=bday, shift=sh)
            db.session.add(sr)

        sr.income = float(s.get("income") or 0.0)
        sr.note = s.get("note") or ""
        sr.is_closed = bool(s.get("is_closed"))
        sr.variable_expense_total = float(s.get("legacy_variable_expense_total") or 0.0)
        sr.fixed_expense_total = float(s.get("legacy_fixed_expense_total") or 0.0)

    db.session.flush()

    expenses = payload.get("expenses") or []
    for e in expenses:
        ds = (e.get("date") or "").strip()
        if ds not in day_map:
            continue
        bday = day_map[ds]
        kind = (e.get("kind") or "").strip().lower()
        catname = (e.get("category_name") or "").strip()
        amount = float(e.get("amount") or 0.0)
        if amount <= 0:
            continue

        cat = get_or_create_cat(kind, catname)
        if not cat:
            continue

        db.session.add(
            ExpenseEntry(
                business_day_id=bday.id,
                kind=kind,
                category_id=cat.id,
                amount=amount,
                note=e.get("note") or "",
            )
        )

    for bday in day_map.values():
        ensure_shifts(bday)
        recalc_day_status(bday)

    db.session.commit()
    return {"imported": imported, "replaced": replaced, "skipped": skipped}


def import_export_excel(filepath: str, mode: str = "skip") -> dict:
    _, ensure_shifts, recalc_day_status, _, _ = _owners()

    wb = openpyxl.load_workbook(filepath, data_only=True)

    imported = 0
    skipped = 0
    replaced = 0

    if "Days" not in wb.sheetnames:
        raise ValueError("El Excel no parece ser un export de PORA (falta hoja 'Days').")

    ws_days = wb["Days"]
    ws_exp = wb["Expenses"] if "Expenses" in wb.sheetnames else None
    ws_cat = wb["Categories"] if "Categories" in wb.sheetnames else None

    if ws_cat:
        for r in range(2, ws_cat.max_row + 1):
            kind = (ws_cat.cell(r, 2).value or "").strip().lower()
            name = (ws_cat.cell(r, 3).value or "").strip()
            if kind not in ("fixed", "variable") or not name:
                continue
            ex = ExpenseCategory.query.filter_by(kind=kind, name=name).first()
            if not ex:
                db.session.add(ExpenseCategory(kind=kind, name=name))
        db.session.flush()

    def get_or_create_cat(kind: str, name: str):
        kind = (kind or "").strip().lower()
        name = re.sub(r"\s+", " ", (name or "").strip())
        if kind not in ("fixed", "variable") or not name:
            return None
        ex = ExpenseCategory.query.filter_by(kind=kind, name=name).first()
        if ex:
            return ex
        ex = ExpenseCategory(kind=kind, name=name)
        db.session.add(ex)
        db.session.flush()
        return ex

    day_map = {}
    for r in range(2, ws_days.max_row + 1):
        ds = (ws_days.cell(r, 1).value or "").strip()
        if not ds:
            continue
        try:
            dd = parse_ymd(ds)
        except Exception:
            continue
        if is_sunday(dd):
            continue

        bday = BusinessDay.query.filter_by(day=dd).first()
        if bday and mode == "skip":
            skipped += 1
            day_map[ds] = bday
            continue

        if not bday:
            bday = BusinessDay(day=dd, note="", status="draft")
            db.session.add(bday)
            db.session.flush()
            ensure_shifts(bday)
            imported += 1
        else:
            replaced += 1

        status = (ws_days.cell(r, 2).value or "draft")
        note = (ws_days.cell(r, 3).value or "")
        income = float(ws_days.cell(r, 4).value or 0.0)
        var_exp = float(ws_days.cell(r, 5).value or 0.0)
        fix_exp = float(ws_days.cell(r, 6).value or 0.0)
        real_profit = ws_days.cell(r, 9).value

        bday.status = str(status)
        bday.note = str(note)
        bday.real_profit = None if real_profit in (None, "") else float(real_profit)

        if mode == "replace":
            ExpenseEntry.query.filter_by(business_day_id=bday.id).delete()
            ShiftRecord.query.filter_by(business_day_id=bday.id).delete()
            db.session.flush()

        ensure_shifts(bday)
        sr_m = ShiftRecord.query.filter_by(business_day_id=bday.id, shift="Mañana").first()
        sr_t = ShiftRecord.query.filter_by(business_day_id=bday.id, shift="Tarde").first()
        if not sr_m:
            sr_m = ShiftRecord(business_day=bday, shift="Mañana")
            db.session.add(sr_m)
        if not sr_t:
            sr_t = ShiftRecord(business_day=bday, shift="Tarde")
            db.session.add(sr_t)

        sr_m.income = income
        sr_m.variable_expense_total = var_exp
        sr_m.fixed_expense_total = fix_exp
        sr_m.is_closed = True
        sr_t.income = 0.0
        sr_t.variable_expense_total = 0.0
        sr_t.fixed_expense_total = 0.0
        sr_t.is_closed = False

        day_map[ds] = bday

    db.session.flush()

    if ws_exp:
        for r in range(2, ws_exp.max_row + 1):
            ds = (ws_exp.cell(r, 1).value or "").strip()
            if ds not in day_map:
                continue
            kind = (ws_exp.cell(r, 2).value or "").strip().lower()
            catname = (ws_exp.cell(r, 3).value or "").strip()
            amount = float(ws_exp.cell(r, 4).value or 0.0)
            note = (ws_exp.cell(r, 5).value or "")
            if amount <= 0:
                continue
            cat = get_or_create_cat(kind, catname)
            if not cat:
                continue
            db.session.add(
                ExpenseEntry(
                    business_day_id=day_map[ds].id,
                    kind=kind,
                    category_id=cat.id,
                    amount=amount,
                    note=str(note),
                )
            )

    for bday in day_map.values():
        ensure_shifts(bday)
        recalc_day_status(bday)

    db.session.commit()
    return {"imported": imported, "replaced": replaced, "skipped": skipped}


@import_export_bp.get("/import/balance")
@login_required
def import_balance_get():
    render_page, _, _, _, _ = _owners()

    body = """
    <h1>Importar Balance Diario</h1>

    <div class="card">
      <form method="post" action="/import/dispatcher" enctype="multipart/form-data" id="importForm">
        <label>Tipo de importación</label>
        <select name="import_type" id="importType">
          <option value="legacy">Importar Balance Diario 2026 (Legacy)</option>
          <option value="export_xlsx">Importar Excel exportado PORA</option>
          <option value="export_json">Importar JSON exportado PORA</option>
        </select>

        <div style="height:12px;"></div>

        <label>Archivo</label>
        <input type="file" name="file" required />

        <div id="legacyBlock" style="margin-top:12px;">
          <label>Hojas a importar (legacy)</label><br/>
          <div style="height:6px;"></div>
          <label><input type="checkbox" name="sheets" value="Enero_26" checked /> Enero_26</label><br/>
          <label><input type="checkbox" name="sheets" value="Febrero_26" checked /> Febrero_26</label>

          <div style="height:12px;"></div>
          <label>Modo (legacy)</label>
          <select name="mode_legacy">
            <option value="skip">No tocar existentes (skip)</option>
            <option value="replace">Reemplazar existentes (replace)</option>
          </select>
        </div>

        <div id="exportBlock" style="margin-top:12px; display:none;">
          <label>Modo (export)</label>
          <select name="mode_export">
            <option value="skip">No tocar existentes (skip)</option>
            <option value="replace">Reemplazar existentes (replace)</option>
          </select>

          <p class="muted" style="margin-top:10px;">
            Excel/JSON exportados: sirven como backup real. En modo "replace" pisa días y gastos del rango importado.
          </p>
        </div>

        <div style="height:12px;"></div>
        <button class="btn primary" type="submit">Importar</button>

        <p class="muted" style="margin-top:10px;">
          Tip: el JSON exportado es el backup más completo (incluye turnos y estructura).
        </p>
      </form>
    </div>

    <script>
      const sel = document.getElementById('importType');
      const legacy = document.getElementById('legacyBlock');
      const ex = document.getElementById('exportBlock');

      function refresh(){
        const v = sel.value;
        if(v === 'legacy'){
          legacy.style.display = 'block';
          ex.style.display = 'none';
        }else{
          legacy.style.display = 'none';
          ex.style.display = 'block';
        }
      }
      sel.addEventListener('change', refresh);
      refresh();
    </script>
    """
    return render_page(body, show_nav=True)


@import_export_bp.post("/import/dispatcher")
@login_required
def import_dispatcher():
    f = request.files.get("file")
    if not f:
        flash("No se recibió archivo.", "error")
        return redirect(url_for("import_export_bp.import_balance_get"))

    import_type = (request.form.get("import_type") or "legacy").strip()

    uploads_dir = os.path.join(INSTANCE_DIR, "uploads")
    os.makedirs(uploads_dir, exist_ok=True)
    save_path = os.path.join(uploads_dir, f.filename)
    f.save(save_path)

    try:
        if import_type == "legacy":
            sheets = request.form.getlist("sheets")
            mode = (request.form.get("mode_legacy") or "skip").strip()
            if not sheets:
                flash("Seleccioná al menos una hoja (Enero_26 / Febrero_26).", "error")
                return redirect(url_for("import_export_bp.import_balance_get"))
            result = import_balance_excel(save_path, sheets, mode=mode)
            flash(
                f"Import LEGACY OK — nuevos: {result['imported']}, reemplazados: {result['replaced']}, salteados: {result['skipped']}",
                "ok",
            )
            return redirect(url_for("dashboard_bp.dashboard_finanzas"))

        mode = (request.form.get("mode_export") or "skip").strip()

        if import_type == "export_json":
            with open(save_path, "r", encoding="utf-8") as fh:
                payload = json.load(fh)
            result = import_export_json(payload, mode=mode)
            flash(
                f"Import JSON OK — nuevos: {result['imported']}, reemplazados: {result['replaced']}, salteados: {result['skipped']}",
                "ok",
            )
            return redirect(url_for("dashboard_bp.dashboard_finanzas"))

        if import_type == "export_xlsx":
            result = import_export_excel(save_path, mode=mode)
            flash(
                f"Import EXCEL OK — nuevos: {result['imported']}, reemplazados: {result['replaced']}, salteados: {result['skipped']}",
                "ok",
            )
            return redirect(url_for("dashboard_bp.dashboard_finanzas"))

        flash("Tipo de importación inválido.", "error")
        return redirect(url_for("import_export_bp.import_balance_get"))

    except Exception as e:
        flash(f"Error importando: {e}", "error")
        return redirect(url_for("import_export_bp.import_balance_get"))


@import_export_bp.get("/api/dashboard")
@login_required
def api_dashboard():
    _, _, _, _, range_series = _owners()

    f = request.args.get("from")
    t = request.args.get("to")
    if not f or not t:
        return jsonify({"error": "params from/to required"}), 400
    d1 = parse_ymd(f)
    d2 = parse_ymd(t)
    series = range_series(d1, d2)
    return jsonify({"from": f, "to": t, "series": series})
